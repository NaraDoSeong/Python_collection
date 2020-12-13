from openpyxl import load_workbook

wb = load_workbook("Samples.xlsx")  # Samples.xlsx 파일을 읽어옴
ws = wb.active
for row in ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=3):
    for r in row:
        print(r.value)
print()
for row in ws.iter_cols(min_row=1, max_row=10, min_col=1, max_col=3):
    for r in row:
        print(r.value)
print()

ws.insert_rows(8) # 8번째 줄 생성
ws.insert_rows(8, 5) # 8번째 줄에 5개 줄 생성

ws.insert_cols(2) # B번째 열 생성
ws.insert_cols(2, 3) # B번째 열에 3개 열 생성

wb.close()