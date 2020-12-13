from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# 데이터 입력
ws.append(['first', 'second', 'third']) # 1줄씩 넣기

for i in range(1, 10):
    ws.append([i, i*10, i*100])

#데이터 추출 1줄씩 (세로)
col_B = ws["B"] # column B의 값 모두 가져옴

for b in col_B:
    print(b.value) # 셀번호의 값을 출력해야 써있는 값이 출력됨
print()
#데이터 추출 2줄이상 (세로)
cols_b_to_c = ws["B:C"] # column B부터 C의 값을 전부 가져옴

for col_b_to_c in cols_b_to_c: # cols_b_to_c=(b전체 cell, c전체 cell)
    for b_to_c in col_b_to_c: # 처음은 col_b_to_c = (b 전체 cell) 그다음 for문이 돌면 col_b_to_c = (c 전체 cell)
        print(b_to_c.value)
print()
#데이터 추출 1줄씩 (가로)
row_1 = ws["1"] # row 1의 값 모두를 한개씩 가져옴

for one in row_1:
    print(one.value) # 셀번호의 값을 출력해야 써있는 값이 출력됨
print()
#데이터 추출 2줄이상 (가로)
rows_1_to_2 = ws["1:2"] # row 1부터 2의 값을 전부 가져옴

for row_1_to_2 in rows_1_to_2:
    for one_to_2 in row_1_to_2:
        print(one_to_2.value)
print()

#셀 번호 추출
row_1 = ws["1"] # row 1의 값 모두를 한개씩 가져옴

#case1
for one in row_1:
    print(one.coordinate) # 셀번호를 출력
print()
#case2
for one in row_1:
    print(one.coordinate)
print()

print(ws.max_column, ws.max_row)
print()


wb.save("Samples.xlsx")
wb.close()
