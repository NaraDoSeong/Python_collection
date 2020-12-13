from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title="TESTS"

ws["A1"] = 2
ws["A2"] = 2

print(ws["A1"].value) # A1의 값
print(ws.cell(1, 1).value) # A1의 값
print(ws.cell(1, 2).value) # B1의 값
print(ws.cell(2, 1).value) # A2의 값
# row = 1, 2, 3, 4 ...
# column = A(1), B(2), C(3), D(4) ...



wb.save("sample.xlsx")
wb.close()